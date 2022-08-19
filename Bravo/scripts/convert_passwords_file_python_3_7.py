#!/bin/python3

# creates key vault yml file
# to run use
# ./export_kv_python_3_7.py -i Keyvault-Names.xlsx -o key-vault.yml

import openpyxl, sys, getopt, yaml

def main(argv):
    workbook = ''
    config = {
      'servers_passwords': [],
    }
    try:
        opts, args = getopt.getopt(argv,"hdi:o:p:",["ifile=","ofile=","passwd="])
    except getopt.GetoptError:
       print('export_passwords_python_3_7.py -i <inputfile> -o <outputfile> [-d]')
       sys.exit(2)
    for opt, arg in opts:
       if opt == '-h':
          print('export_passwords_python_3_7.py -i <inputfile> -o <outputfile> [-d]')
       elif opt in ("-i", "--ifile"):
          workbook = arg
       elif opt in ("-o", "--ofile"):
          outputfile = arg
       elif opt in ("-p", "--passwd"):
          xpass = arg
       elif opt in ("-d", "--delete"):
          delete = True

    print('Opening workbook "', workbook)
    wb = openpyxl.load_workbook(workbook,data_only=True)
    wb.guest_types = True
    sheet = wb.active
#    sheet.protection.password = 'bravo'

    for i in range (2, sheet.max_row + 1, 1):

         Hostname   = str(sheet.cell(row = i, column =  1).value)
         IPaddress  = str(sheet.cell(row = i, column =  2).value)
         Username   = str(sheet.cell(row = i, column =  3).value)
         Password   = str(sheet.cell(row = i, column =  4).value)

         val = {
                 'name':     Hostname,
                 'ip':       IPaddress,
                'user':     Username,
                 'password': Password
               }

         config['servers_passwords'].append(val)

    print('Writing to ' + outputfile)
    stream = open(outputfile,'w')
    yaml.dump(config, stream, allow_unicode=False, tags=False, default_flow_style=False, canonical=False, encoding='ascii' )

if __name__ == "__main__":
   main(sys.argv[1:])

